import pytest
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.alert import Alert

@pytest.fixture(scope="module")
def driver():
    # Setup
    driver = webdriver.Chrome()
    driver.implicitly_wait(10)
    yield driver
    # Teardown
    driver.quit()

def test_verify_title(driver):
    driver.get("https://www.saucedemo.com/")
    assert "Swag Labs" in driver.title

def read_excel_testdata(testcaseName):
    final_list = []
    workbook=openpyxl.load_workbook("C:\Users\Neeraja\Documents\CAPSTONE PROJECT-NEERAJA\Test Data.xlsx")
    sheet = workbook['TestData']
    total_rows=sheet.max_row
    total_cols=sheet.max_column
    rows=iter(sheet) 
    firstRow=next(rows)
    cell=iter(firstRow)
    k=0
    column=0
    while next(cell):
        cellvalue=next(cell)
        if cellvalue.value=="Testcases" :
            column=k
        k+=1
    print(column)

    while next(rows):
     r=next(rows)
     colvalue=r[column]
     if colvalue.value==testcaseName :
         cv=iter(r)
         while next(cv) :
             final_list.append(next(cv).value)
         
    return final_list

 
    for r in range(2,total_rows+1):
        row_list = []
        for c in range(2,total_cols+1):
         row_list.append(sheet.cell(r,c).value)
        final_list.append(row_list)
    return final_list



# @pytest.mark.parametrize("username,password",read_excel_testdata(testcaseName))
def test_successfulLogin(username,password,driver,testcaseName):
    testcaseName="Login positive"
    username=driver.find_element_by_xpath("//*[@placeholder='Username']")
    password=driver.find_element_by_xpath("//*[@placeholder='Password']")
    login=driver.find_element_by_xpath("//input[@type='submit']")
    username.send_keys(read_excel_testdata(testcaseName).get(1))
    password.send_keys(read_excel_testdata(testcaseName).get(2))
    login.click
    alert= Alert(driver)
    alert.accept()
    test_verify_title(driver)

def test_unsuccessfulLogin(username,password,driver,testcaseName):
    testcaseName="Login negative"
    username=driver.find_element_by_xpath("//*[@placeholder='Username']")
    password=driver.find_element_by_xpath("//*[@placeholder='Password']")
    login=driver.find_element_by_xpath("//input[@type='submit']")
    username.send_keys(read_excel_testdata(testcaseName).get(1))
    password.send_keys(read_excel_testdata(testcaseName).get(2))
    login.click
    errormessage=driver.find_element_by_xpath("//h3[@data-test='error']")
    actualmessage=errormessage.text
    expectedmessage="Epic sadface: Username and password do not match any user in this service"
    assert actualmessage==expectedmessage
    




